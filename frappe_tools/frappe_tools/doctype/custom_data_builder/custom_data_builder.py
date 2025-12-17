# Copyright (c) 2025, sakthi123msd@gmail.com and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import requests
from io import BytesIO
from bs4 import BeautifulSoup

from frappe.www.printview import get_print_style
from frappe.utils import get_site_path
import base64
from essdee.essdee.doctype.sd_signature.sd_signature import get_user_signature



base_template = """

<!DOCTYPE html>
<html>
<head>
	<meta charset="utf-8">

	<style type="text/css">
		body {
			margin: 0;
			padding: 0;
			font-family: Arial, Helvetica, sans-serif;
			font-size: 12px;
			line-height: 1.5;
			color: #000000;
		}

		table {
			border-collapse: collapse;
			width: 100%;
		}

		td {
			vertical-align: top;
		}

		.page {
			padding: 30px;
		}

		.text {
			margin-bottom: 10px;
			text-align: justify;
		}

		.bold {
			font-weight: bold;
		}

		.mt-5 { margin-top: 5px; }
		.mt-10 { margin-top: 10px; }
		.mt-15 { margin-top: 15px; }
		.mt-20 { margin-top: 20px; }

		.signature {
			margin-top: 40px;
		}

		.footer {
			margin-top: 50px;
			font-size: 10px;
			color: #555555;
		}
	</style>

	<style type="text/css">
		{{ css }}
	</style>

</head>

<body>
	<div class="page">
		{% if header %}
			<div class="header">
				{{ header }}
			</div>
		{% endif %}

		<!-- MAIN BODY CONTENT -->
		{{ body }}

		{% if footer %}
		<div class="footer">
			{{ footer }}
		</div>
		{% endif %}

	</div>
</body>
</html>


"""

class CustomDataBuilder(Document):

	# def validate(self):
	# 	if self.print_format:
	# 		if not frappe.db.get_value("Print Format", self.print_format, 'custom_format'):
	# 			frappe.throw("Can't use not custom print formats")
	
	def before_submit(self):
		self.submitted_user = frappe.session.user
		# if not self.print_format:
		# 	frappe.throw("Select print format before submit")

@frappe.whitelist()
def get_document_uploaded_values(doc_name, limit=10):
	resource = frappe.db.get_value("Custom Data Builder", doc_name, "resource_document")
	if not resource:
		frappe.throw("No file attached.")

	columns, rows = load_excel_file(resource)
	limit = int(limit)

	data_dicts = []
	for row in rows[: int(limit) ]:
		item = {}
		for idx, col in enumerate(columns):
			item[col] = row[idx] if idx < len(row) else None
		data_dicts.append(item)

	return {
		"data": data_dicts,
		"total_count": len(rows)
	}

def download_file_from_url(url):
	resp = requests.get(url)
	if resp.status_code != 200:
		frappe.throw(f"Unable to download file. Status {resp.status_code}")
	return resp.content


def read_excel_from_bytes(file_bytes):
	
	try:
		import openpyxl
		wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
		sheet = wb.active

		columns = []
		rows = []

		for cell in sheet[1]:
			columns.append(cell.value)

		for row in sheet.iter_rows(min_row=2, values_only=True):
			rows.append(list(row))

		return columns, rows

	except Exception:
		import xlrd
		workbook = xlrd.open_workbook(file_contents=file_bytes)
		sheet = workbook.sheet_by_index(0)

		columns = sheet.row_values(0)
		rows = []

		for r in range(1, sheet.nrows):
			rows.append(sheet.row_values(r))

		return columns, rows

def load_excel_file(resource):

	if resource.startswith("/files/") or resource.startswith("/private/files/"):
		file_path = frappe.get_site_path(resource.lstrip("/"))
		with open(file_path, "rb") as f:
			file_bytes = f.read()
		return read_excel_from_bytes(file_bytes)

	if resource.startswith("http://") or resource.startswith("https://"):
		file_bytes = download_file_from_url(resource)
		return read_excel_from_bytes(file_bytes)

	if frappe.db.exists("File", resource):
		file_doc = frappe.get_doc("File", resource)
		if file_doc.is_private:
			file_path = frappe.get_site_path("private", "files", file_doc.file_name)
		else:
			file_path = frappe.get_site_path("public", "files", file_doc.file_name)

		with open(file_path, "rb") as f:
			file_bytes = f.read()

		return read_excel_from_bytes(file_bytes)

	frappe.throw("Unsupported file format or file not found.")


@frappe.whitelist()
def get_list_details(doctype, filters=None, limit=10):
	filters = frappe.json.loads(filters or "[]")

	data = frappe.get_all(
		doctype,
		filters=filters,
		limit=limit,
		fields="*"
	)

	total_count = frappe.db.count(doctype, filters)

	return {
		"data": data,
		"total_count": total_count
	}


@frappe.whitelist()
def download_excel(name):
	from openpyxl import Workbook
	doc = frappe.get_doc("Data Builder", name)

	if doc.resource_document_type == "Doctype":
		filters = frappe.parse_json(doc.filter_json or "[]")

		data = frappe.get_all(
			doc._doctype,
			filters=filters,
			fields="*",
		)
	else:
		data = frappe.parse_json(doc.extracted_data_json or "[]")

	if not data:
		frappe.throw("No data available to export.")

	wb = Workbook()
	ws = wb.active
	ws.title = "Data"

	headers = list(data[0].keys())
	ws.append(headers)

	for row in data:
		ws.append([row.get(h, "") for h in headers])

	file_name = f"{doc.name}.xlsx"
	file_path = frappe.get_site_path("private", "files", file_name)
	wb.save(file_path)

	frappe.local.response.filecontent = open(file_path, "rb").read()
	frappe.local.response.type = "download"
	frappe.local.response.filename = file_name


@frappe.whitelist()
def send_email(doc):
	"""
	Trigger background job to send emails extracted from Excel file.
	"""
	if not doc:
		frappe.throw("Docname is required.")
	
	builder = frappe.get_doc("Custom Data Builder", doc)
	
	if not builder.email_account:
		frappe.throw("Email Account is required.")

	if not builder.target_field:
		frappe.throw("Target field (column name) is required.")

	frappe.enqueue(
		method=process_email_sending,
		doc=doc,
		user = frappe.session.user,
		queue="long",
		timeout=600
	)
	# process_email_sending(doc=doc, user = frappe.session.user)


def get_doctype_rows(builder):
	filters = frappe.parse_json(builder.filter_json or "[]")

	rows = frappe.get_all(
		builder._doctype,      
		filters=filters,
		fields="*"
	)
	
	# Convert to list of dicts/Strings that are serializable
	return [dict(row) for row in rows]

def get_excel_rows(builder):
	if not builder.resource_document:
		frappe.throw("No Excel file uploaded.")
		
	columns, excel_rows = load_excel_file(builder.resource_document)

	rows = []
	for row in excel_rows:
		item = {}
		for idx, col in enumerate(columns):
			item[col] = row[idx] if idx < len(row) else None
		rows.append(item)

	return rows

@frappe.whitelist()
def get_preview_content(doc):
	builder = frappe.get_doc("Custom Data Builder", doc)
	
	row_data = get_processed_row_data(builder, limit=1)
	if not row_data:
		frappe.throw("No data found in uploaded file to generate preview.")
	
	first_row = row_data[0]
	
	# Render Email
	email_subject = frappe.render_template(builder.subject, first_row)
	email_body = frappe.render_template(builder.main_body_content, first_row)
	
	attachment_url = None
	if builder.add_attachment and builder.print_format:
		pdf_content = generate_pdf_attachment(builder, first_row)
		if pdf_content:
			fname = f"Data_Builder_Preview_{builder.name}.pdf"
			saved_file = frappe.get_doc({
				"doctype": "File",
				"file_name": fname,
				"content": pdf_content,
				"is_private": 1
			})
			saved_file.save(ignore_permissions=True)
			attachment_url = saved_file.file_url

	return {
		"subject": email_subject,
		"html_body": email_body,
		"attachment_url": attachment_url
	}

def get_processed_row_data(builder, limit=None):
	data_source = builder.resource_document_type
	if data_source == "Doctype":
		rows = get_doctype_rows(builder)
	else:
		rows = get_excel_rows(builder)
		
	if limit and len(rows) > limit:
		rows = rows[:limit]

	# Fetch Global Constants
	global_constants = {}
	if builder.global_constants:
		for item in builder.global_constants:
			if item.key:
				global_constants[item.key] = item.value

	processed_rows = []
	for i in rows:
		row_data = i.copy()
		row_data.update(global_constants)
		processed_rows.append(row_data)
		
	return processed_rows

def generate_pdf_attachment(builder, row_data):
	import datetime
	if not builder.print_format:
		return None

	print_format_details = frappe.get_doc("Print Format", builder.print_format)
	print_format_html = print_format_details.html
	print_style = get_print_style(print_format=print_format_details)
	
	header_html = None
	footer_html = None
	row_data['sd_approved_by'] = builder.submitted_user
	row_data['posting_date'] = datetime.datetime.now().strftime("%Y-%m-%d")
	
	if builder.letter_head:
		lh = frappe.get_doc("Letter Head", builder.letter_head)
		header_html = lh.content
		footer_html = lh.footer

	attachment_html_rendered = frappe.render_template(
		print_format_html,
		context={
			"data" : row_data
		},
	)
	
	full_attachment_html = frappe.render_template(base_template, {
		"header": header_html,
		"footer": footer_html,
		"body": attachment_html_rendered, 
		"css": print_style
	})
	full_attachment_html = image_conversion(full_attachment_html)
	from frappe.utils.pdf import get_pdf
	return get_pdf(full_attachment_html)


def process_email_sending(doc, user=None):
	if not user:
		user = frappe.session.user
	
	builder = frappe.get_doc("Custom Data Builder", doc)
	
	# Use common data fetching
	rows = get_processed_row_data(builder)

	processable_docs = []
	from_email_account = frappe.get_doc("Email Account", builder.email_account)
	
	for row_data in rows:
		# Create a log entry
		log_doc = frappe.new_doc("Data Builder Share Log")
		log_doc.update({
			"data_builder" : builder.name,
			"log_status" : "Pending",
			"user" : user,
			"share_type" : "Email",
			"print_format" : builder.print_format if builder.add_attachment else None,
			"args" : frappe.json.dumps({
				"from_email" : from_email_account.email_id,
				"to" : row_data.get(builder.target_field),
				"cc" : builder.cc,
				"bcc" : builder.bcc,
				"subject" : builder.subject
			}),
			"format_data" : frappe.json.dumps(row_data)
		})
		log_doc.flags.ignore_permissions = True
		log_doc.submit()
		frappe.db.commit()
		processable_docs.append(log_doc.name)
	
	for i in processable_docs:
		try:
			document = frappe.get_doc("Data Builder Share Log", i)
			args = frappe.json.loads(document.args)
			row_data = frappe.json.loads(document.format_data)
			
			email_subject = frappe.render_template(args.get('subject'), row_data)
			email_body = frappe.render_template(builder.main_body_content, row_data)

			attachments = []
		
			if builder.add_attachment and builder.print_format:
				pdf_content = generate_pdf_attachment(builder, row_data)
				if pdf_content:
					if document.attachment_naming_field:
						fname = row_data.get(document.attachment_naming_field, builder.name)
						attachments.append({
							"fname": f"{fname}.pdf",
							"fcontent": pdf_content
						})

			email_queue = frappe.sendmail(
				sender=args.get('from_email'),
				recipients=args.get('to'),
				cc=args.get('cc'),
				bcc=args.get('bcc'),
				subject=email_subject,
				message=email_body,
				attachments=attachments,
				now=True
			)
			frappe.db.set_value("Data Builder Share Log", i, 'email_queue', email_queue.name)
			frappe.db.set_value("Data Builder Share Log", i, 'log_status', 'Processing')
		except Exception as e:
			err = frappe.log_error("Data Builder Share Error")
			frappe.db.set_value("Data Builder Share Log", i, 'error_log', err.name)
			frappe.db.set_value("Data Builder Share Log", i, 'log_status', 'Failed')
		finally :
			frappe.db.commit()

@frappe.whitelist()
def delete_old_previews():
	from frappe.utils import now_datetime
	from datetime import timedelta
	
	cutoff_time = now_datetime() - timedelta(minutes=15)
	
	files = frappe.get_all("File", filters={
		"file_name": ["like", "Data_Builder_Preview_%"],
		"creation": ["<", cutoff_time],
		"is_private": 1
	}, fields=["name", "file_name"])
	
	for file_doc in files:
		try:
			frappe.delete_doc("File", file_doc.name, ignore_permissions=True)
		except Exception:
			frappe.log_error(f"Failed to delete preview file {file_doc.file_name}", "Preview Cleanup Error")
	
	if files:
		frappe.db.commit()
	

@frappe.whitelist()
def poll_update_status_processing_data_share():

	lists = frappe.get_list("Data Builder Share Log", filters = [
		['log_status', '=', 'Processing'],
		['share_type', '=', 'Email']
	], fields = ['email_queue', 'name'])

	for i in lists:
		# Check if queue exists
		if not frappe.db.exists("Email Queue", i['email_queue']):
			continue

		status = frappe.get_value("Email Queue", i['email_queue'], 'status')
		if status in ['Not Sent', 'Sending']:
			continue

		if status == 'Sent':
			frappe.db.set_value("Data Builder Share Log", i['name'] , 'log_status', 'Success')
		else:
			frappe.db.set_value("Data Builder Share Log",i['name'], 'log_status' ,'Failed' )

def image_conversion(html):
	soup = BeautifulSoup(html)
	for img in soup.find_all('img'):
		img_url = img.get('src')
		img_path = img_url.split('/')
		if img_path[1] == "private":
			img_url = get_site_path(*img_path)
		elif img_path[1] == "files":
			img_url = get_site_path("public", *img_path)
		else:
			continue
		img_url = img_url.replace(" ", "%20")
		ext = img_url.split('.')[-1]
		encoded = base64.b64encode(open(img_url, 'rb').read()).decode('utf-8')
		prefix = f'data:image/{ext};base64,'
		img.attrs['src'] = prefix + encoded
	return (str(soup))
